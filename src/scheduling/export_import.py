"""
BiScheduler Schedule Export/Import
Venezuelan K12 schedule data exchange functionality
Support for Excel and CSV formats matching Venezuelan standards
"""

import csv
import io
import json
from datetime import datetime, timezone
from typing import Dict, List, Optional, BinaryIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.scheduling.services import ScheduleManager
from src.models.tenant import DayOfWeek


class VenezuelanScheduleExporter:
    """
    Export schedules in Venezuelan K12 standard formats
    Supports both student schedules and teacher workload reports
    """

    def __init__(self, schedule_manager: ScheduleManager):
        self.schedule_manager = schedule_manager
        self.academic_year = schedule_manager.academic_year

    def export_student_schedule_excel(self, section_id: int) -> bytes:
        """
        Export student schedule in Venezuelan format Excel

        Args:
            section_id: Section to export

        Returns:
            Excel file as bytes
        """
        session = self.schedule_manager.SessionLocal()

        try:
            from src.models.tenant import Section, TimePeriod, ScheduleAssignment

            # Get section information
            section = session.query(Section).filter_by(id=section_id).first()
            if not section:
                raise ValueError(f"Section {section_id} not found")

            # Get time periods
            time_periods = session.query(TimePeriod).filter_by(
                academic_year=self.academic_year
            ).order_by(TimePeriod.display_order).all()

            # Get all assignments for section
            assignments = session.query(ScheduleAssignment).filter_by(
                section_id=section_id,
                academic_year=self.academic_year,
                is_active=True
            ).all()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Horario {section.name}"

            # Venezuelan standard styles
            header_font = Font(name='Arial', size=12, bold=True)
            cell_font = Font(name='Arial', size=10)
            header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = f"HORARIO DE CLASES - {section.name.upper()}"
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')

            # Subtitle
            ws.merge_cells('A2:G2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = f"AÑO ACADÉMICO {self.academic_year}"
            subtitle_cell.font = Font(name='Arial', size=12)
            subtitle_cell.alignment = Alignment(horizontal='center')

            # Headers
            headers = ['HORA', 'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            # Organization assignments by day and time
            schedule_grid = {}
            for assignment in assignments:
                day = assignment.day_of_week.value
                period = assignment.time_period.period_name

                if period not in schedule_grid:
                    schedule_grid[period] = {}

                schedule_grid[period][day] = {
                    'subject': assignment.subject.subject_name,
                    'teacher': assignment.teacher.teacher_name,
                    'classroom': assignment.classroom.name
                }

            # Fill schedule data
            row = 5
            for period in time_periods:
                period_name = period.period_name

                # Time column
                time_cell = ws.cell(row=row, column=1)
                time_cell.value = f"{period.start_time.strftime('%H:%M')}\n{period.end_time.strftime('%H:%M')}"
                time_cell.font = cell_font
                time_cell.alignment = Alignment(horizontal='center', vertical='center')
                time_cell.border = border

                # Day columns
                day_mapping = {
                    'lunes': 2, 'martes': 3, 'miercoles': 4, 'jueves': 5, 'viernes': 6
                }

                for day_spanish, col in day_mapping.items():
                    cell = ws.cell(row=row, column=col)

                    if period_name in schedule_grid and day_spanish in schedule_grid[period_name]:
                        assignment_data = schedule_grid[period_name][day_spanish]

                        if period.is_break:
                            cell.value = "RECREO"
                            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                        else:
                            # Venezuelan format: MATERIA\nPROFESOR\n(Aula X)
                            cell.value = f"{assignment_data['subject']}\n{assignment_data['teacher']}\n({assignment_data['classroom']})"
                    else:
                        if period.is_break:
                            cell.value = "RECREO"
                            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                        else:
                            cell.value = ""

                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border

                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            for col in range(2, 7):
                ws.column_dimensions[get_column_letter(col)].width = 18

            # Adjust row heights
            for row_num in range(5, row):
                ws.row_dimensions[row_num].height = 60

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        finally:
            session.close()

    def export_teacher_workload_excel(self) -> bytes:
        """
        Export teacher workload report in Venezuelan CARGA HORARIA format

        Returns:
            Excel file as bytes
        """
        session = self.schedule_manager.SessionLocal()

        try:
            from src.models.tenant import Teacher, TeacherSubject, Subject
            from sqlalchemy import func

            # Get all teachers with their subjects and hours
            teachers_data = session.query(
                Teacher.id,
                Teacher.teacher_name,
                Teacher.cedula,
                Teacher.current_weekly_hours,
                Teacher.max_weekly_hours
            ).filter_by(
                academic_year=self.academic_year,
                is_active=True
            ).order_by(Teacher.teacher_name).all()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "CARGA HORARIA"

            # Venezuelan CARGA HORARIA styles
            header_font = Font(name='Arial', size=12, bold=True)
            cell_font = Font(name='Arial', size=10)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = f"CARGA HORARIA DE PROFESORES - AÑO ACADÉMICO {self.academic_year}"
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')

            # Headers (Venezuelan standard format)
            headers = ['N°', 'NOMBRES Y APELLIDOS', 'CÉDULA', 'ASIGNATURA', 'CARGA HORARIA']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            # Fill teacher data
            row = 4
            for i, (teacher_id, name, cedula, current_hours, max_hours) in enumerate(teachers_data, 1):
                # Get subjects for this teacher
                teacher_subjects = session.query(
                    Subject.subject_name,
                    TeacherSubject.weekly_hours
                ).join(TeacherSubject).filter(
                    TeacherSubject.teacher_id == teacher_id,
                    TeacherSubject.academic_year == self.academic_year,
                    TeacherSubject.is_active == True
                ).all()

                # Combine subjects
                subjects_text = ", ".join([
                    f"{subject_name} ({hours}h)"
                    for subject_name, hours in teacher_subjects
                ])

                # Fill row data
                ws.cell(row=row, column=1).value = i
                ws.cell(row=row, column=2).value = name
                ws.cell(row=row, column=3).value = cedula or "N/A"
                ws.cell(row=row, column=4).value = subjects_text
                ws.cell(row=row, column=5).value = current_hours or 0

                # Apply formatting
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.border = border
                    if col in [1, 5]:  # Number and hours columns
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 4:  # Subjects column
                        cell.alignment = Alignment(wrap_text=True)

                row += 1

            # Summary row
            ws.cell(row=row + 1, column=1).value = "TOTAL"
            ws.cell(row=row + 1, column=1).font = header_font
            ws.cell(row=row + 1, column=2).value = f"{len(teachers_data)} profesores"
            ws.cell(row=row + 1, column=2).font = header_font

            # Column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 12

            # Row heights for subject cells
            for row_num in range(4, row):
                ws.row_dimensions[row_num].height = 30

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        finally:
            session.close()

    def export_schedule_csv(self, section_id: int) -> str:
        """
        Export schedule as CSV for data exchange

        Args:
            section_id: Section to export

        Returns:
            CSV content as string
        """
        session = self.schedule_manager.SessionLocal()

        try:
            from src.models.tenant import ScheduleAssignment

            assignments = session.query(ScheduleAssignment).filter_by(
                section_id=section_id,
                academic_year=self.academic_year,
                is_active=True
            ).all()

            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            writer.writerow([
                'section_id', 'section_name', 'day_of_week', 'period_name',
                'start_time', 'end_time', 'subject_name', 'teacher_name',
                'classroom_name', 'academic_year'
            ])

            # Data rows
            for assignment in assignments:
                writer.writerow([
                    assignment.section_id,
                    assignment.section.name,
                    assignment.day_of_week.value,
                    assignment.time_period.period_name,
                    assignment.time_period.start_time.strftime('%H:%M:%S'),
                    assignment.time_period.end_time.strftime('%H:%M:%S'),
                    assignment.subject.subject_name,
                    assignment.teacher.teacher_name,
                    assignment.classroom.name,
                    assignment.academic_year
                ])

            return output.getvalue()

        finally:
            session.close()


class VenezuelanScheduleImporter:
    """
    Import schedules from Venezuelan K12 standard formats
    Supports Excel and CSV import with validation
    """

    def __init__(self, schedule_manager: ScheduleManager):
        self.schedule_manager = schedule_manager
        self.academic_year = schedule_manager.academic_year

    def import_from_csv(self, csv_content: str, created_by: str = None) -> Dict[str, any]:
        """
        Import schedule assignments from CSV

        Args:
            csv_content: CSV content string
            created_by: User who initiated the import

        Returns:
            Import results
        """
        try:
            reader = csv.DictReader(io.StringIO(csv_content))

            imported_count = 0
            failed_count = 0
            errors = []

            for row_num, row in enumerate(reader, 1):
                try:
                    # Map CSV data to assignment parameters
                    day_of_week = DayOfWeek(row['day_of_week'])

                    # Get entity IDs from names (you might need to implement lookup functions)
                    teacher_id = self._get_teacher_id_by_name(row['teacher_name'])
                    subject_id = self._get_subject_id_by_name(row['subject_name'])
                    section_id = self._get_section_id_by_name(row['section_name'])
                    classroom_id = self._get_classroom_id_by_name(row['classroom_name'])
                    time_period_id = self._get_time_period_id_by_name(row['period_name'])

                    if not all([teacher_id, subject_id, section_id, classroom_id, time_period_id]):
                        errors.append(f"Row {row_num}: Missing entity references")
                        failed_count += 1
                        continue

                    # Create assignment
                    result = self.schedule_manager.create_schedule_assignment(
                        teacher_id=teacher_id,
                        subject_id=subject_id,
                        section_id=section_id,
                        classroom_id=classroom_id,
                        time_period_id=time_period_id,
                        day_of_week=day_of_week,
                        created_by=created_by,
                        validate_conflicts=True
                    )

                    if result['status'] == 'success':
                        imported_count += 1
                    else:
                        errors.append(f"Row {row_num}: {result['message']}")
                        failed_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    failed_count += 1

            return {
                'status': 'success',
                'imported_assignments': imported_count,
                'failed_assignments': failed_count,
                'errors': errors
            }

        except Exception as e:
            return {
                'status': 'error',
                'message': f'Import failed: {str(e)}'
            }

    def _get_teacher_id_by_name(self, teacher_name: str) -> Optional[int]:
        """Get teacher ID by name"""
        session = self.schedule_manager.SessionLocal()
        try:
            from src.models.tenant import Teacher
            teacher = session.query(Teacher).filter_by(
                teacher_name=teacher_name,
                academic_year=self.academic_year
            ).first()
            return teacher.id if teacher else None
        finally:
            session.close()

    def _get_subject_id_by_name(self, subject_name: str) -> Optional[int]:
        """Get subject ID by name"""
        session = self.schedule_manager.SessionLocal()
        try:
            from src.models.tenant import Subject
            subject = session.query(Subject).filter_by(
                subject_name=subject_name,
                academic_year=self.academic_year
            ).first()
            return subject.id if subject else None
        finally:
            session.close()

    def _get_section_id_by_name(self, section_name: str) -> Optional[int]:
        """Get section ID by name"""
        session = self.schedule_manager.SessionLocal()
        try:
            from src.models.tenant import Section
            section = session.query(Section).filter_by(
                name=section_name,
                academic_year=self.academic_year
            ).first()
            return section.id if section else None
        finally:
            session.close()

    def _get_classroom_id_by_name(self, classroom_name: str) -> Optional[int]:
        """Get classroom ID by name"""
        session = self.schedule_manager.SessionLocal()
        try:
            from src.models.tenant import Classroom
            classroom = session.query(Classroom).filter_by(
                name=classroom_name,
                is_active=True
            ).first()
            return classroom.id if classroom else None
        finally:
            session.close()

    def _get_time_period_id_by_name(self, period_name: str) -> Optional[int]:
        """Get time period ID by name"""
        session = self.schedule_manager.SessionLocal()
        try:
            from src.models.tenant import TimePeriod
            period = session.query(TimePeriod).filter_by(
                period_name=period_name,
                academic_year=self.academic_year
            ).first()
            return period.id if period else None
        finally:
            session.close()


def create_schedule_template_excel() -> bytes:
    """
    Create an empty Venezuelan schedule template for manual data entry

    Returns:
        Excel template as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Horario"

    # Venezuelan template styles
    header_font = Font(name='Arial', size=12, bold=True)
    header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "PLANTILLA DE HORARIO - BISCHEDULER"
    title_cell.font = Font(name='Arial', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Instructions
    ws.merge_cells('A2:G2')
    instructions = ws['A2']
    instructions.value = "Complete este formato con la información de clases para importar al sistema"
    instructions.font = Font(name='Arial', size=10, italic=True)
    instructions.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['HORA', 'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'OBSERVACIONES']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Sample time periods (Venezuelan standard)
    time_periods = [
        "07:00-07:40", "07:40-08:20", "08:20-09:00", "09:00-09:40",
        "09:40-10:00", "10:00-10:40", "10:40-11:20", "11:20-12:00",
        "12:00-12:40", "12:40-13:20", "13:20-14:00", "14:00-14:20"
    ]

    for i, period in enumerate(time_periods, 5):
        ws.cell(row=i, column=1).value = period
        ws.cell(row=i, column=1).font = Font(name='Arial', size=10)
        ws.cell(row=i, column=1).border = border
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')

        # Add borders to other columns
        for col in range(2, 8):
            cell = ws.cell(row=i, column=col)
            cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['G'].width = 20

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()