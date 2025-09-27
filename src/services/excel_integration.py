"""
Excel Integration Service
Phase 3: Venezuelan K12 Excel Import/Export System
Handles teachers, students, classrooms, and subjects data processing
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook
import os
import tempfile
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
import json
import logging
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from sqlalchemy.orm import Session
from src.models.tenant import (
    Teacher, Student, Classroom, Subject, Section,
    TimePeriod, ScheduleAssignment, AcademicPeriod
)

logger = logging.getLogger(__name__)


class ExcelValidationError(Exception):
    """Custom exception for Excel validation errors"""
    pass


class ExcelIntegrationService:
    """
    Venezuelan K12 Excel Integration Service
    Handles import/export of educational data with government format compliance
    """

    def __init__(self, db_session: Session, tenant_id: str):
        self.db_session = db_session
        self.tenant_id = tenant_id
        self.allowed_extensions = {'.xlsx', '.xls'}
        self.max_file_size = 10 * 1024 * 1024  # 10MB

        # Venezuelan K12 validation rules
        self.grade_levels = ['1er año', '2do año', '3er año', '4to año', '5to año']
        self.section_types = ['A', 'B', 'C', 'D']
        self.venezuelan_subjects = [
            'CASTELLANO Y LITERATURA', 'INGLÉS', 'MATEMÁTICAS', 'FÍSICA', 'QUÍMICA',
            'BIOLOGÍA', 'GHC PARA LA SOBERANIA NACIONAL', 'EDUCACIÓN FÍSICA',
            'ORIENTACIÓN VOCACIONAL', 'LOGICA MATEMÁTICA', 'ARTE Y PATRIMONIO'
        ]

    # ============================================================================
    # FILE UPLOAD AND VALIDATION
    # ============================================================================

    def validate_uploaded_file(self, file: FileStorage) -> bool:
        """
        Validate uploaded Excel file
        Returns True if valid, raises ExcelValidationError if invalid
        """
        if not file:
            raise ExcelValidationError("No file provided")

        if not file.filename:
            raise ExcelValidationError("No filename provided")

        filename = secure_filename(file.filename)
        file_ext = os.path.splitext(filename)[1].lower()

        if file_ext not in self.allowed_extensions:
            raise ExcelValidationError(f"Invalid file type. Allowed: {', '.join(self.allowed_extensions)}")

        # Check file size (rough estimate from stream)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning

        if file_size > self.max_file_size:
            raise ExcelValidationError(f"File too large. Maximum size: {self.max_file_size // 1024 // 1024}MB")

        return True

    def save_uploaded_file(self, file: FileStorage) -> str:
        """Save uploaded file to temporary location and return path"""
        self.validate_uploaded_file(file)

        # Create temporary file
        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"upload_{self.tenant_id}_{timestamp}_{secure_filename(file.filename)}"
        temp_path = os.path.join(temp_dir, filename)

        file.save(temp_path)
        logger.info(f"File saved to temporary location: {temp_path}")

        return temp_path

    # ============================================================================
    # TEACHERS IMPORT/EXPORT
    # ============================================================================

    def import_teachers_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Import teachers from Excel file
        Expected columns: teacher_name, teacher_email, specialization, phone, cedula
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            logger.info(f"Loading teachers from {file_path}, found {len(df)} rows")

            # Validate required columns
            required_columns = ['teacher_name', 'specialization']
            optional_columns = ['teacher_email', 'phone', 'cedula', 'department']

            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ExcelValidationError(f"Missing required columns: {missing_columns}")

            results = {
                'success': 0,
                'errors': 0,
                'duplicates': 0,
                'details': [],
                'imported_teachers': []
            }

            for index, row in df.iterrows():
                try:
                    # Extract teacher data
                    teacher_name = str(row['teacher_name']).strip()
                    specialization = str(row['specialization']).strip()

                    if not teacher_name or teacher_name.lower() in ['nan', 'none']:
                        results['errors'] += 1
                        results['details'].append(f"Row {index + 2}: Missing teacher name")
                        continue

                    # Check for duplicates
                    existing_teacher = self.db_session.query(Teacher).filter(
                        Teacher.teacher_name == teacher_name
                    ).first()

                    if existing_teacher:
                        results['duplicates'] += 1
                        results['details'].append(f"Row {index + 2}: Teacher {teacher_name} already exists")
                        continue

                    # Create new teacher
                    teacher = Teacher(
                        teacher_name=teacher_name,
                        specialization=specialization,
                        teacher_email=str(row.get('teacher_email', '')).strip() if pd.notna(row.get('teacher_email')) else None,
                        phone=str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else None,
                        cedula=str(row.get('cedula', '')).strip() if pd.notna(row.get('cedula')) else None,
                        department=str(row.get('department', '')).strip() if pd.notna(row.get('department')) else None,
                        is_active=True
                    )

                    self.db_session.add(teacher)
                    results['success'] += 1
                    results['imported_teachers'].append(teacher_name)
                    results['details'].append(f"Row {index + 2}: Successfully imported {teacher_name}")

                except Exception as e:
                    results['errors'] += 1
                    results['details'].append(f"Row {index + 2}: Error - {str(e)}")
                    logger.error(f"Error importing teacher row {index + 2}: {e}")

            if results['success'] > 0:
                self.db_session.commit()
                logger.info(f"Successfully imported {results['success']} teachers")
            else:
                self.db_session.rollback()
                logger.warning("No teachers imported, rolling back transaction")

            return results

        except Exception as e:
            self.db_session.rollback()
            logger.error(f"Failed to import teachers: {e}")
            raise ExcelValidationError(f"Import failed: {str(e)}")

    def export_teachers_to_excel(self) -> str:
        """Export all teachers to Excel file"""
        try:
            teachers = self.db_session.query(Teacher).filter(Teacher.is_active == True).all()

            # Prepare data
            data = []
            for teacher in teachers:
                data.append({
                    'ID': teacher.id,
                    'Nombre del Docente': teacher.teacher_name,
                    'Especialización': teacher.specialization,
                    'Email': teacher.teacher_email or '',
                    'Teléfono': teacher.phone or '',
                    'Cédula': teacher.cedula or '',
                    'Departamento': teacher.department or '',
                    'Estado': 'Activo' if teacher.is_active else 'Inactivo',
                    'Fecha de Registro': teacher.created_at.strftime('%Y-%m-%d') if teacher.created_at else ''
                })

            df = pd.DataFrame(data)

            # Create temporary file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            temp_path = os.path.join(tempfile.gettempdir(), f"teachers_export_{self.tenant_id}_{timestamp}.xlsx")

            # Export with formatting
            with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Docentes', index=False)

                # Format the worksheet
                worksheet = writer.sheets['Docentes']

                # Header formatting
                header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(f"Teachers exported to {temp_path}")
            return temp_path

        except Exception as e:
            logger.error(f"Failed to export teachers: {e}")
            raise ExcelValidationError(f"Export failed: {str(e)}")

    # ============================================================================
    # STUDENTS IMPORT/EXPORT
    # ============================================================================

    def import_students_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Import students from Excel file
        Expected columns: student_name, cedula, grade_level, section_name, gender
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            logger.info(f"Loading students from {file_path}, found {len(df)} rows")

            # Validate required columns
            required_columns = ['student_name', 'cedula', 'grade_level', 'section_name']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ExcelValidationError(f"Missing required columns: {missing_columns}")

            results = {
                'success': 0,
                'errors': 0,
                'duplicates': 0,
                'details': [],
                'imported_students': []
            }

            for index, row in df.iterrows():
                try:
                    # Extract student data
                    student_name = str(row['student_name']).strip()
                    cedula = str(row['cedula']).strip()
                    grade_level = str(row['grade_level']).strip()
                    section_name = str(row['section_name']).strip()

                    if not student_name or student_name.lower() in ['nan', 'none']:
                        results['errors'] += 1
                        results['details'].append(f"Row {index + 2}: Missing student name")
                        continue

                    if not cedula or cedula.lower() in ['nan', 'none']:
                        results['errors'] += 1
                        results['details'].append(f"Row {index + 2}: Missing cedula")
                        continue

                    # Check for duplicates
                    existing_student = self.db_session.query(Student).filter(
                        Student.cedula == cedula
                    ).first()

                    if existing_student:
                        results['duplicates'] += 1
                        results['details'].append(f"Row {index + 2}: Student with cedula {cedula} already exists")
                        continue

                    # Find or create section
                    section = self.db_session.query(Section).filter(
                        Section.name == section_name,
                        Section.grade_level == grade_level
                    ).first()

                    if not section:
                        # Create new section
                        section = Section(
                            name=section_name,
                            grade_level=grade_level,
                            capacity=35,  # Default Venezuelan section capacity
                            is_active=True
                        )
                        self.db_session.add(section)
                        self.db_session.flush()  # Get section ID

                    # Create new student
                    student = Student(
                        student_name=student_name,
                        cedula=cedula,
                        section_id=section.id,
                        gender=str(row.get('gender', '')).strip() if pd.notna(row.get('gender')) else None,
                        birth_date=pd.to_datetime(row['birth_date']).date() if pd.notna(row.get('birth_date')) else None,
                        parent_name=str(row.get('parent_name', '')).strip() if pd.notna(row.get('parent_name')) else None,
                        parent_phone=str(row.get('parent_phone', '')).strip() if pd.notna(row.get('parent_phone')) else None,
                        is_active=True
                    )

                    self.db_session.add(student)
                    results['success'] += 1
                    results['imported_students'].append(student_name)
                    results['details'].append(f"Row {index + 2}: Successfully imported {student_name}")

                except Exception as e:
                    results['errors'] += 1
                    results['details'].append(f"Row {index + 2}: Error - {str(e)}")
                    logger.error(f"Error importing student row {index + 2}: {e}")

            if results['success'] > 0:
                self.db_session.commit()
                logger.info(f"Successfully imported {results['success']} students")
            else:
                self.db_session.rollback()
                logger.warning("No students imported, rolling back transaction")

            return results

        except Exception as e:
            self.db_session.rollback()
            logger.error(f"Failed to import students: {e}")
            raise ExcelValidationError(f"Import failed: {str(e)}")

    # ============================================================================
    # CLASSROOMS IMPORT/EXPORT
    # ============================================================================

    def import_classrooms_from_excel(self, file_path: str) -> Dict[str, Any]:
        """Import classrooms from Excel file"""
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            logger.info(f"Loading classrooms from {file_path}, found {len(df)} rows")

            required_columns = ['classroom_name', 'capacity']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ExcelValidationError(f"Missing required columns: {missing_columns}")

            results = {
                'success': 0,
                'errors': 0,
                'duplicates': 0,
                'details': [],
                'imported_classrooms': []
            }

            for index, row in df.iterrows():
                try:
                    classroom_name = str(row['classroom_name']).strip()
                    capacity = int(row['capacity']) if pd.notna(row['capacity']) else 35

                    if not classroom_name or classroom_name.lower() in ['nan', 'none']:
                        results['errors'] += 1
                        results['details'].append(f"Row {index + 2}: Missing classroom name")
                        continue

                    # Check for duplicates
                    existing_classroom = self.db_session.query(Classroom).filter(
                        Classroom.classroom_name == classroom_name
                    ).first()

                    if existing_classroom:
                        results['duplicates'] += 1
                        results['details'].append(f"Row {index + 2}: Classroom {classroom_name} already exists")
                        continue

                    # Create new classroom
                    classroom = Classroom(
                        classroom_name=classroom_name,
                        capacity=capacity,
                        classroom_type=str(row.get('classroom_type', 'Regular')).strip(),
                        location=str(row.get('location', '')).strip() if pd.notna(row.get('location')) else None,
                        equipment=str(row.get('equipment', '')).strip() if pd.notna(row.get('equipment')) else None,
                        is_active=True
                    )

                    self.db_session.add(classroom)
                    results['success'] += 1
                    results['imported_classrooms'].append(classroom_name)
                    results['details'].append(f"Row {index + 2}: Successfully imported {classroom_name}")

                except Exception as e:
                    results['errors'] += 1
                    results['details'].append(f"Row {index + 2}: Error - {str(e)}")
                    logger.error(f"Error importing classroom row {index + 2}: {e}")

            if results['success'] > 0:
                self.db_session.commit()
                logger.info(f"Successfully imported {results['success']} classrooms")
            else:
                self.db_session.rollback()

            return results

        except Exception as e:
            self.db_session.rollback()
            logger.error(f"Failed to import classrooms: {e}")
            raise ExcelValidationError(f"Import failed: {str(e)}")

    # ============================================================================
    # SCHEDULE EXPORT
    # ============================================================================

    def export_schedule_to_excel(self, academic_period_id: Optional[int] = None) -> str:
        """Export complete schedule to Excel file with Venezuelan format"""
        try:
            # Get academic period
            if academic_period_id:
                academic_period = self.db_session.query(AcademicPeriod).get(academic_period_id)
            else:
                academic_period = self.db_session.query(AcademicPeriod).filter(
                    AcademicPeriod.is_active == True
                ).first()

            if not academic_period:
                raise ExcelValidationError("No active academic period found")

            # Get all schedule assignments
            assignments = self.db_session.query(ScheduleAssignment).filter(
                ScheduleAssignment.academic_period_id == academic_period.id
            ).all()

            # Get time periods
            time_periods = self.db_session.query(TimePeriod).order_by(TimePeriod.start_time).all()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Horario General"

            # Create schedule matrix
            days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

            # Headers
            ws['A1'] = 'Período'
            ws['B1'] = 'Hora'
            for col, day in enumerate(days, start=3):
                ws.cell(row=1, column=col, value=day)

            # Time periods and schedule data
            row = 2
            for period in time_periods:
                ws.cell(row=row, column=1, value=f"Período {period.period_number}")
                ws.cell(row=row, column=2, value=f"{period.start_time.strftime('%H:%M')} - {period.end_time.strftime('%H:%M')}")

                for col, day in enumerate(days, start=3):
                    # Find assignment for this time/day
                    day_enum = self._get_day_enum(day)
                    assignment = next((a for a in assignments
                                     if a.time_period_id == period.id and a.day_of_week == day_enum), None)

                    if assignment:
                        cell_value = f"{assignment.subject.subject_name}\n{assignment.teacher.teacher_name}\n{assignment.classroom.classroom_name}"
                        ws.cell(row=row, column=col, value=cell_value)
                    else:
                        ws.cell(row=row, column=col, value="")

                row += 1

            # Format the worksheet
            self._format_schedule_worksheet(ws)

            # Create temporary file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            temp_path = os.path.join(tempfile.gettempdir(), f"schedule_export_{self.tenant_id}_{timestamp}.xlsx")

            wb.save(temp_path)
            logger.info(f"Schedule exported to {temp_path}")

            return temp_path

        except Exception as e:
            logger.error(f"Failed to export schedule: {e}")
            raise ExcelValidationError(f"Export failed: {str(e)}")

    def _get_day_enum(self, day_name: str):
        """Convert Spanish day name to enum"""
        day_mapping = {
            'Lunes': 'monday',
            'Martes': 'tuesday',
            'Miércoles': 'wednesday',
            'Jueves': 'thursday',
            'Viernes': 'friday'
        }
        return day_mapping.get(day_name, 'monday')

    def _format_schedule_worksheet(self, worksheet):
        """Apply Venezuelan educational formatting to schedule worksheet"""
        # Header formatting
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # ============================================================================
    # EXCEL TEMPLATE GENERATION
    # ============================================================================

    def create_teachers_template(self) -> str:
        """Create Excel template for teachers import"""
        data = {
            'teacher_name': ['MARIA NIETO', 'FLORMAR HERNANDEZ', 'STEFANY ROMERO'],
            'specialization': ['Matemáticas', 'Castellano', 'Inglés'],
            'teacher_email': ['maria.nieto@example.com', 'flormar.h@example.com', 'stefany.r@example.com'],
            'phone': ['0414-1234567', '0424-2345678', '0412-3456789'],
            'cedula': ['12345678', '23456789', '34567890'],
            'department': ['Ciencias', 'Humanidades', 'Idiomas']
        }

        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(tempfile.gettempdir(), f"teachers_template_{timestamp}.xlsx")

        df.to_excel(temp_path, index=False, sheet_name='Docentes')
        logger.info(f"Teachers template created: {temp_path}")

        return temp_path

    def create_students_template(self) -> str:
        """Create Excel template for students import"""
        data = {
            'student_name': ['Juan Pérez', 'María González', 'Carlos Rodríguez'],
            'cedula': ['30123456', '29234567', '28345678'],
            'grade_level': ['1er año', '2do año', '3er año'],
            'section_name': ['A', 'B', 'A'],
            'gender': ['M', 'F', 'M'],
            'birth_date': ['2008-05-15', '2007-08-20', '2006-12-10'],
            'parent_name': ['Pedro Pérez', 'Ana González', 'Luis Rodríguez'],
            'parent_phone': ['0414-1111111', '0424-2222222', '0412-3333333']
        }

        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(tempfile.gettempdir(), f"students_template_{timestamp}.xlsx")

        df.to_excel(temp_path, index=False, sheet_name='Estudiantes')
        logger.info(f"Students template created: {temp_path}")

        return temp_path

    def create_classrooms_template(self) -> str:
        """Create Excel template for classrooms import"""
        data = {
            'classroom_name': ['Aula 1', 'Aula 2', 'Lab Física', 'Cancha 1'],
            'capacity': [35, 35, 25, 50],
            'classroom_type': ['Regular', 'Regular', 'Laboratorio', 'Deportes'],
            'location': ['Planta Baja', 'Planta Baja', 'Segundo Piso', 'Exterior'],
            'equipment': ['Pizarra, Proyector', 'Pizarra', 'Equipos de laboratorio', 'Equipos deportivos']
        }

        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(tempfile.gettempdir(), f"classrooms_template_{timestamp}.xlsx")

        df.to_excel(temp_path, index=False, sheet_name='Aulas')
        logger.info(f"Classrooms template created: {temp_path}")

        return temp_path

    # ============================================================================
    # UTILITY METHODS
    # ============================================================================

    def cleanup_temp_file(self, file_path: str):
        """Clean up temporary file"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Cleaned up temporary file: {file_path}")
        except Exception as e:
            logger.warning(f"Failed to cleanup temp file {file_path}: {e}")

    def get_import_statistics(self) -> Dict[str, int]:
        """Get current database statistics"""
        return {
            'teachers': self.db_session.query(Teacher).filter(Teacher.is_active == True).count(),
            'students': self.db_session.query(Student).filter(Student.is_active == True).count(),
            'classrooms': self.db_session.query(Classroom).filter(Classroom.is_active == True).count(),
            'subjects': self.db_session.query(Subject).filter(Subject.is_active == True).count(),
            'sections': self.db_session.query(Section).filter(Section.is_active == True).count()
        }